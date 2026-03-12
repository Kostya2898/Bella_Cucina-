from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, abort
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_wtf import FlaskForm
from flask_wtf.csrf import CSRFProtect
from flask_migrate import Migrate
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_mail import Mail, Message
from wtforms import StringField, PasswordField, SubmitField, IntegerField, FloatField, TextAreaField, SelectField
from wtforms.fields import DateTimeLocalField
from wtforms.validators import DataRequired, Email, EqualTo, ValidationError, Length, Optional, NumberRange
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from functools import wraps
from threading import Thread
import os
import qrcode
import io
import base64
import logging
import re
import json
import openpyxl
import hashlib
import hmac
from openpyxl.styles import Font, PatternFill, Alignment

# ─── Додаток ──────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.config['SECRET_KEY']                   = os.environ.get('SECRET_KEY', 'bella_cucina_secret_2024_ЗМІНІТЬ_У_ПРОДАКШНІ')
app.config['SQLALCHEMY_DATABASE_URI']      = os.environ.get('DATABASE_URL', 'sqlite:///bella_cucina.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SESSION_COOKIE_HTTPONLY']      = True       # JS не може читати cookie
app.config['SESSION_COOKIE_SAMESITE']      = 'Lax'     # захист від CSRF
app.config['PERMANENT_SESSION_LIFETIME']   = timedelta(hours=2)  # сесія 2 години
app.config['WTF_CSRF_TIME_LIMIT']          = 3600       # CSRF токен діє 1 годину

# ─── Flask-Mail ───────────────────────────────────────────────────────────────
app.config['MAIL_SERVER']        = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT']          = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS']       = os.environ.get('MAIL_USE_TLS', 'true').lower() == 'true'
app.config['MAIL_USERNAME']      = os.environ.get('MAIL_USERNAME', '')   # ваш gmail
app.config['MAIL_PASSWORD']      = os.environ.get('MAIL_PASSWORD', '')   # app password
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', 'Bella Cucina <noreply@bellacucina.com>')
MAIL_ENABLED = bool(os.environ.get('MAIL_USERNAME'))  # email лише якщо налаштовано

# ─── Stripe ───────────────────────────────────────────────────────────────────
STRIPE_PUBLIC_KEY  = os.environ.get('STRIPE_PUBLIC_KEY', '')
STRIPE_SECRET_KEY  = os.environ.get('STRIPE_SECRET_KEY', '')
STRIPE_WEBHOOK_SECRET = os.environ.get('STRIPE_WEBHOOK_SECRET', '')
STRIPE_ENABLED     = bool(STRIPE_PUBLIC_KEY and STRIPE_SECRET_KEY)

# ─── LiqPay ───────────────────────────────────────────────────────────────────
LIQPAY_PUBLIC_KEY  = os.environ.get('LIQPAY_PUBLIC_KEY', '')
LIQPAY_PRIVATE_KEY = os.environ.get('LIQPAY_PRIVATE_KEY', '')
LIQPAY_ENABLED     = bool(LIQPAY_PUBLIC_KEY and LIQPAY_PRIVATE_KEY)

db      = SQLAlchemy(app)
migrate = Migrate(app, db)
csrf    = CSRFProtect(app)
mail    = Mail(app)

login_manager                      = LoginManager(app)
login_manager.login_view           = 'login'
login_manager.login_message        = 'Будь ласка, увійдіть для доступу до цієї сторінки.'
login_manager.login_message_category = 'info'

# Rate limiter — обмеження кількості запитів
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["300 per hour", "60 per minute"],
    storage_uri="memory://"
)

# ─── Логування ────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler('security.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
security_logger = logging.getLogger('security')


def log_security_event(event: str, details: str = ''):
    ip = request.remote_addr if request else 'N/A'
    user = current_user.email if current_user and current_user.is_authenticated else 'anonymous'
    security_logger.warning(f'[{event}] user={user} ip={ip} {details}')


# ─── Email сповіщення ─────────────────────────────────────────────────────────

def send_email_async(app, msg):
    """Надсилає email у фоновому потоці, щоб не блокувати запит."""
    with app.app_context():
        try:
            mail.send(msg)
        except Exception as e:
            security_logger.error(f'[EMAIL_ERROR] {e}')


def send_email(subject: str, recipients: list, html_body: str):
    """Відправляє HTML-email асинхронно. Нічого не робить якщо пошта не налаштована."""
    if not MAIL_ENABLED:
        return
    msg = Message(subject=subject, recipients=recipients, html=html_body)
    Thread(target=send_email_async, args=(app, msg), daemon=True).start()


def send_order_confirmation(user, order):
    items_html = ''.join(
        f'<tr><td style="padding:6px 12px;">{item.dish.name}</td>'
        f'<td style="padding:6px 12px; text-align:center;">{item.quantity}</td>'
        f'<td style="padding:6px 12px; text-align:right;">${item.price * item.quantity:.2f}</td></tr>'
        for item in order.items
    )
    discount_row = (
        f'<tr><td colspan="2" style="padding:4px 12px; color:#4caf50;">Знижка ({order.discount:.0f}%):</td>'
        f'<td style="padding:4px 12px; text-align:right; color:#4caf50;">−${(sum(i.price*i.quantity for i in order.items) - order.total_price):.2f}</td></tr>'
        if order.discount else ''
    )
    html = f"""
    <div style="font-family:sans-serif;max-width:560px;margin:0 auto;background:#0a0e27;color:#e0e0e0;border-radius:12px;overflow:hidden;">
      <div style="background:linear-gradient(135deg,#d4af37,#f0e68c);padding:24px;text-align:center;">
        <h1 style="margin:0;color:#1a1a1a;font-size:1.6rem;">✦ Bella Cucina ✦</h1>
        <p style="margin:6px 0 0;color:#333;font-size:0.95rem;">Замовлення підтверджено!</p>
      </div>
      <div style="padding:24px;">
        <p>Привіт, <strong>{user.username}</strong>! 👋</p>
        <p>Ваше замовлення <strong>№{order.id}</strong> успішно оформлено. Дякуємо!</p>
        <table style="width:100%;border-collapse:collapse;margin:16px 0;background:rgba(255,255,255,0.05);border-radius:8px;">
          <thead>
            <tr style="background:rgba(212,175,55,0.2);">
              <th style="padding:8px 12px;text-align:left;color:#d4af37;">Страва</th>
              <th style="padding:8px 12px;text-align:center;color:#d4af37;">К-ть</th>
              <th style="padding:8px 12px;text-align:right;color:#d4af37;">Сума</th>
            </tr>
          </thead>
          <tbody>{items_html}</tbody>
          <tfoot>
            {discount_row}
            <tr style="border-top:1px solid rgba(212,175,55,0.3);">
              <td colspan="2" style="padding:8px 12px;font-weight:bold;color:#f0e68c;">Разом:</td>
              <td style="padding:8px 12px;text-align:right;font-weight:bold;color:#d4af37;font-size:1.1rem;">${order.total_price:.2f}</td>
            </tr>
          </tfoot>
        </table>
        <p style="color:#aaa;font-size:0.9rem;">Очікуйте на підтвердження від нашого персоналу. Зазвичай це займає до 10 хвилин.</p>
        <div style="text-align:center;margin-top:24px;">
          <a href="{request.host_url}order/{order.id}"
             style="background:linear-gradient(135deg,#d4af37,#f0e68c);color:#1a1a1a;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:bold;">
            Переглянути замовлення
          </a>
        </div>
      </div>
      <div style="padding:16px;text-align:center;color:#555;font-size:0.82rem;border-top:1px solid rgba(255,255,255,0.05);">
        © 2026 Bella Cucina · вул. Хрещатик, 1, Київ
      </div>
    </div>"""
    send_email(f'Замовлення №{order.id} підтверджено — Bella Cucina', [user.email], html)


def send_order_status_update(user, order):
    status_labels = {
        'confirmed':  ('✅ Підтверджено',  '#4caf50'),
        'preparing':  ('👨‍🍳 Готується',    '#ff9800'),
        'ready':      ('🔔 Готово',        '#2196f3'),
        'delivered':  ('🚀 Доставлено',    '#4caf50'),
        'cancelled':  ('❌ Скасовано',     '#f44336'),
    }
    label, color = status_labels.get(order.status, (order.status, '#aaa'))
    html = f"""
    <div style="font-family:sans-serif;max-width:560px;margin:0 auto;background:#0a0e27;color:#e0e0e0;border-radius:12px;overflow:hidden;">
      <div style="background:linear-gradient(135deg,#d4af37,#f0e68c);padding:24px;text-align:center;">
        <h1 style="margin:0;color:#1a1a1a;font-size:1.6rem;">✦ Bella Cucina ✦</h1>
      </div>
      <div style="padding:24px;">
        <p>Привіт, <strong>{user.username}</strong>!</p>
        <p>Статус вашого замовлення <strong>№{order.id}</strong> змінився:</p>
        <div style="text-align:center;margin:24px 0;">
          <span style="background:rgba(255,255,255,0.08);padding:14px 32px;border-radius:30px;font-size:1.2rem;font-weight:bold;color:{color};">
            {label}
          </span>
        </div>
        <div style="text-align:center;margin-top:24px;">
          <a href="{request.host_url}order/{order.id}"
             style="background:linear-gradient(135deg,#d4af37,#f0e68c);color:#1a1a1a;padding:12px 28px;border-radius:8px;text-decoration:none;font-weight:bold;">
            Деталі замовлення
          </a>
        </div>
      </div>
      <div style="padding:16px;text-align:center;color:#555;font-size:0.82rem;border-top:1px solid rgba(255,255,255,0.05);">
        © 2026 Bella Cucina · вул. Хрещатик, 1, Київ
      </div>
    </div>"""
    send_email(f'Замовлення №{order.id}: {label} — Bella Cucina', [user.email], html)


def send_booking_confirmation(user, booking):
    html = f"""
    <div style="font-family:sans-serif;max-width:560px;margin:0 auto;background:#0a0e27;color:#e0e0e0;border-radius:12px;overflow:hidden;">
      <div style="background:linear-gradient(135deg,#d4af37,#f0e68c);padding:24px;text-align:center;">
        <h1 style="margin:0;color:#1a1a1a;font-size:1.6rem;">✦ Bella Cucina ✦</h1>
        <p style="margin:6px 0 0;color:#333;font-size:0.95rem;">Бронювання отримано!</p>
      </div>
      <div style="padding:24px;">
        <p>Привіт, <strong>{user.username}</strong>! 🍽️</p>
        <p>Ми отримали ваш запит на бронювання столика. Очікуйте підтвердження.</p>
        <table style="width:100%;border-collapse:collapse;margin:16px 0;background:rgba(255,255,255,0.05);border-radius:8px;">
          <tr>
            <td style="padding:10px 14px;color:#aaa;">📅 Дата і час:</td>
            <td style="padding:10px 14px;color:#f0e68c;font-weight:bold;">{booking.booking_date.strftime('%d.%m.%Y о %H:%M')}</td>
          </tr>
          <tr style="background:rgba(255,255,255,0.03);">
            <td style="padding:10px 14px;color:#aaa;">👥 Кількість гостей:</td>
            <td style="padding:10px 14px;font-weight:bold;">{booking.guests} ос.</td>
          </tr>
          {'<tr><td style="padding:10px 14px;color:#aaa;">📝 Побажання:</td><td style="padding:10px 14px;">' + booking.notes + '</td></tr>' if booking.notes else ''}
        </table>
        <p style="color:#aaa;font-size:0.9rem;">Після підтвердження ви отримаєте ще один лист з деталями.</p>
      </div>
      <div style="padding:16px;text-align:center;color:#555;font-size:0.82rem;border-top:1px solid rgba(255,255,255,0.05);">
        © 2026 Bella Cucina · вул. Хрещатик, 1, Київ
      </div>
    </div>"""
    send_email('Бронювання отримано — Bella Cucina', [user.email], html)


def send_booking_status_update(user, booking):
    status_labels = {
        'confirmed':  ('✅ Підтверджено', '#4caf50'),
        'cancelled':  ('❌ Скасовано',    '#f44336'),
    }
    label, color = status_labels.get(booking.status, (booking.status, '#aaa'))
    html = f"""
    <div style="font-family:sans-serif;max-width:560px;margin:0 auto;background:#0a0e27;color:#e0e0e0;border-radius:12px;overflow:hidden;">
      <div style="background:linear-gradient(135deg,#d4af37,#f0e68c);padding:24px;text-align:center;">
        <h1 style="margin:0;color:#1a1a1a;font-size:1.6rem;">✦ Bella Cucina ✦</h1>
      </div>
      <div style="padding:24px;">
        <p>Привіт, <strong>{user.username}</strong>!</p>
        <p>Статус вашого бронювання на <strong>{booking.booking_date.strftime('%d.%m.%Y о %H:%M')}</strong> змінився:</p>
        <div style="text-align:center;margin:24px 0;">
          <span style="background:rgba(255,255,255,0.08);padding:14px 32px;border-radius:30px;font-size:1.2rem;font-weight:bold;color:{color};">
            {label}
          </span>
        </div>
        {'<p style="text-align:center;color:#4caf50;">Чекаємо на вас! 🍷</p>' if booking.status == 'confirmed' else ''}
      </div>
      <div style="padding:16px;text-align:center;color:#555;font-size:0.82rem;border-top:1px solid rgba(255,255,255,0.05);">
        © 2026 Bella Cucina · вул. Хрещатик, 1, Київ
      </div>
    </div>"""
    send_email(f'Бронювання {label} — Bella Cucina', [user.email], html)


# ─── LiqPay хелпери ───────────────────────────────────────────────────────────

def liqpay_encode(data: dict) -> tuple[str, str]:
    """Повертає (data_b64, signature) для LiqPay форми."""
    import base64, json
    data_str  = json.dumps(data)
    data_b64  = base64.b64encode(data_str.encode()).decode()
    raw_sig   = LIQPAY_PRIVATE_KEY + data_b64 + LIQPAY_PRIVATE_KEY
    signature = base64.b64encode(hashlib.sha1(raw_sig.encode()).digest()).decode()
    return data_b64, signature


def liqpay_decode(data_b64: str, received_sig: str) -> dict | None:
    """Перевіряє підпис і повертає декодований dict або None."""
    import base64, json
    raw_sig   = LIQPAY_PRIVATE_KEY + data_b64 + LIQPAY_PRIVATE_KEY
    expected  = base64.b64encode(hashlib.sha1(raw_sig.encode()).digest()).decode()
    if not hmac.compare_digest(expected, received_sig):
        return None
    return json.loads(base64.b64decode(data_b64).decode())


# ─── Stripe хелпер ────────────────────────────────────────────────────────────

def get_stripe():
    """Ліниво ініціалізує stripe лише якщо налаштовано."""
    if not STRIPE_ENABLED:
        return None
    try:
        import stripe
        stripe.api_key = STRIPE_SECRET_KEY
        return stripe
    except ImportError:
        return None


# ─── Захисні HTTP-заголовки ───────────────────────────────────────────────────

@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options']    = 'nosniff'
    response.headers['X-Frame-Options']           = 'SAMEORIGIN'
    response.headers['X-XSS-Protection']          = '1; mode=block'
    response.headers['Referrer-Policy']           = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy']        = 'geolocation=(), microphone=(), camera=()'
    response.headers['Content-Security-Policy']   = (
        "default-src 'self'; "
        "style-src 'self' 'unsafe-inline'; "
        "script-src 'self' 'unsafe-inline'; "
        "img-src 'self' data: https:; "
        "font-src 'self';"
    )
    return response


# ─── Захист від брутфорсу ─────────────────────────────────────────────────────

# Зберігаємо спроби входу в пам'яті: {ip: {'count': N, 'locked_until': datetime}}
login_attempts = {}
MAX_ATTEMPTS   = 5
LOCKOUT_MIN    = 15


def is_ip_locked(ip: str) -> bool:
    data = login_attempts.get(ip)
    if not data:
        return False
    if data.get('locked_until') and datetime.utcnow() < data['locked_until']:
        return True
    if data.get('locked_until') and datetime.utcnow() >= data['locked_until']:
        login_attempts.pop(ip, None)
    return False


def record_failed_login(ip: str):
    if ip not in login_attempts:
        login_attempts[ip] = {'count': 0, 'locked_until': None}
    login_attempts[ip]['count'] += 1
    if login_attempts[ip]['count'] >= MAX_ATTEMPTS:
        login_attempts[ip]['locked_until'] = datetime.utcnow() + timedelta(minutes=LOCKOUT_MIN)
        log_security_event('BRUTE_FORCE_LOCKED', f'ip={ip}')


def reset_login_attempts(ip: str):
    login_attempts.pop(ip, None)


def get_remaining_lockout(ip: str) -> int:
    data = login_attempts.get(ip)
    if data and data.get('locked_until'):
        delta = data['locked_until'] - datetime.utcnow()
        return max(0, int(delta.total_seconds() // 60))
    return 0


# ─── Санітизація вводу ────────────────────────────────────────────────────────

def sanitize_string(value: str, max_length: int = 500) -> str:
    if not value:
        return ''
    value = value.strip()[:max_length]
    # Видаляємо небезпечні HTML теги
    value = re.sub(r'<[^>]*>', '', value)
    return value


# ─── Декоратори ───────────────────────────────────────────────────────────────

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            log_security_event('UNAUTHORIZED_ADMIN_ACCESS', f'path={request.path}')
            flash('Потрібен доступ адміністратора', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated


# ─── Моделі ───────────────────────────────────────────────────────────────────

class User(UserMixin, db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    username      = db.Column(db.String(80), unique=True, nullable=False)
    email         = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin      = db.Column(db.Boolean, default=False)
    is_blocked    = db.Column(db.Boolean, default=False)   # блокування акаунту
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    orders        = db.relationship('Order',   backref='customer', lazy=True, cascade='all, delete-orphan')
    bookings      = db.relationship('Booking', backref='customer', lazy=True, cascade='all, delete-orphan')
    reviews       = db.relationship('Review',  backref='author',   lazy=True, cascade='all, delete-orphan')

    def set_password(self, password):
        self.password_hash = generate_password_hash(password, method='pbkdf2:sha256', salt_length=16)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def is_password_strong(password: str) -> bool:
        return len(password) >= 6


class Dish(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    name        = db.Column(db.String(120), nullable=False)
    description = db.Column(db.Text)
    price       = db.Column(db.Float, nullable=False)
    category    = db.Column(db.String(50), nullable=False)
    image_url   = db.Column(db.String(255))
    is_active   = db.Column(db.Boolean, default=True)
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)
    order_items = db.relationship('OrderItem', backref='dish', lazy=True, cascade='all, delete-orphan')
    reviews     = db.relationship('Review',    backref='dish', lazy=True, cascade='all, delete-orphan')

    @property
    def avg_rating(self):
        if not self.reviews:
            return 0
        return round(sum(r.rating for r in self.reviews) / len(self.reviews), 1)

    @property
    def review_count(self):
        return len(self.reviews)


class Order(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    user_id     = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    total_price = db.Column(db.Float, nullable=False)
    status      = db.Column(db.String(20), default='pending')
    promo_code  = db.Column(db.String(50))
    discount    = db.Column(db.Float, default=0.0)
    created_at  = db.Column(db.DateTime, default=datetime.utcnow)
    items       = db.relationship('OrderItem', backref='order', lazy=True, cascade='all, delete-orphan')


class OrderItem(db.Model):
    id       = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey('order.id'), nullable=False)
    dish_id  = db.Column(db.Integer, db.ForeignKey('dish.id'), nullable=False)
    quantity = db.Column(db.Integer, nullable=False)
    price    = db.Column(db.Float, nullable=False)


class Booking(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    user_id      = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    booking_date = db.Column(db.DateTime, nullable=False)
    guests       = db.Column(db.Integer, nullable=False)
    table_number = db.Column(db.Integer)
    status       = db.Column(db.String(20), default='pending')
    notes        = db.Column(db.Text)
    created_at   = db.Column(db.DateTime, default=datetime.utcnow)


class Review(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    user_id    = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    dish_id    = db.Column(db.Integer, db.ForeignKey('dish.id'), nullable=False)
    rating     = db.Column(db.Integer, nullable=False)
    comment    = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class PromoCode(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    code       = db.Column(db.String(50), unique=True, nullable=False)
    discount   = db.Column(db.Float, nullable=False)
    is_active  = db.Column(db.Boolean, default=True)
    uses_left  = db.Column(db.Integer, default=None)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class Payment(db.Model):
    id             = db.Column(db.Integer, primary_key=True)
    order_id       = db.Column(db.Integer, db.ForeignKey('order.id'), nullable=False)
    user_id        = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    amount         = db.Column(db.Float, nullable=False)
    currency       = db.Column(db.String(10), default='UAH')
    provider       = db.Column(db.String(20), nullable=False)   # 'stripe' | 'liqpay' | 'cash'
    status         = db.Column(db.String(20), default='pending')  # pending|paid|refunded|failed
    external_id    = db.Column(db.String(255))   # Stripe PaymentIntent id / LiqPay order_id
    refund_reason  = db.Column(db.Text)
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at     = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    order          = db.relationship('Order', backref=db.backref('payment', uselist=False))
    payer          = db.relationship('User',  backref='payments')


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ─── Jinja2 фільтри ───────────────────────────────────────────────────────────

@app.template_filter('fromjson')
def fromjson_filter(s):
    return json.loads(s)


app.jinja_env.globals['enumerate'] = enumerate


# ─── Форми ────────────────────────────────────────────────────────────────────

class RegistrationForm(FlaskForm):
    username         = StringField("Ім'я користувача", validators=[DataRequired(), Length(min=3, max=80)])
    email            = StringField('Email', validators=[DataRequired(), Email()])
    password         = PasswordField('Пароль', validators=[DataRequired(), Length(min=6)])
    confirm_password = PasswordField('Підтвердіть пароль', validators=[DataRequired(), EqualTo('password')])
    submit           = SubmitField('Зареєструватися')

    def validate_username(self, username):
        clean = sanitize_string(username.data)
        if not re.match(r'^[a-zA-Zа-яА-ЯіІїЇєЄ0-9_]+$', clean):
            raise ValidationError("Лише літери, цифри та символ _")
        if User.query.filter_by(username=clean).first():
            raise ValidationError("Це ім'я вже зайняте")

    def validate_email(self, email):
        if User.query.filter_by(email=email.data.lower().strip()).first():
            raise ValidationError('Цей email вже використовується')


class LoginForm(FlaskForm):
    email    = StringField('Email', validators=[DataRequired(), Email()])
    password = PasswordField('Пароль', validators=[DataRequired()])
    submit   = SubmitField('Увійти')


class ProfileForm(FlaskForm):
    username = StringField("Ім'я користувача", validators=[DataRequired(), Length(min=3, max=80)])
    email    = StringField('Email', validators=[DataRequired(), Email()])
    submit   = SubmitField('Зберегти зміни')


class ChangePasswordForm(FlaskForm):
    current_password = PasswordField('Поточний пароль', validators=[DataRequired()])
    new_password     = PasswordField('Новий пароль', validators=[DataRequired(), Length(min=6)])
    confirm_password = PasswordField('Підтвердіть новий пароль', validators=[DataRequired(), EqualTo('new_password')])
    submit           = SubmitField('Змінити пароль')


class ReviewForm(FlaskForm):
    rating  = SelectField('Оцінка', choices=[('5','⭐⭐⭐⭐⭐'),('4','⭐⭐⭐⭐'),('3','⭐⭐⭐'),('2','⭐⭐'),('1','⭐')], validators=[DataRequired()])
    comment = TextAreaField('Коментар', validators=[Optional(), Length(max=500)])
    submit  = SubmitField('Залишити відгук')


CATEGORIES = [
    ('appetizers', 'Закуски'),
    ('main',       'Основні страви'),
    ('desserts',   'Десерти'),
    ('drinks',     'Напої'),
]


class DishForm(FlaskForm):
    name        = StringField('Назва страви', validators=[DataRequired(), Length(max=120)])
    description = TextAreaField('Опис', validators=[Optional(), Length(max=1000)])
    price       = FloatField('Ціна', validators=[DataRequired(), NumberRange(min=0.01, max=10000)])
    category    = SelectField('Категорія', choices=CATEGORIES)
    image_url   = StringField('URL зображення', validators=[Optional(), Length(max=255)])
    submit      = SubmitField('Додати страву')


class BookingForm(FlaskForm):
    booking_date = DateTimeLocalField('Дата та час', format='%Y-%m-%dT%H:%M', validators=[DataRequired()])
    guests       = IntegerField('Кількість гостей', validators=[DataRequired(), NumberRange(min=1, max=20)])
    notes        = TextAreaField('Побажання', validators=[Optional(), Length(max=500)])
    submit       = SubmitField('Забронювати столик')


class PromoForm(FlaskForm):
    code      = StringField('Код знижки', validators=[DataRequired(), Length(min=3, max=50)])
    discount  = FloatField('Знижка (%)', validators=[DataRequired(), NumberRange(min=1, max=100)])
    uses_left = IntegerField('Кількість використань (0 = необмежено)', validators=[Optional()])
    submit    = SubmitField('Створити промокод')


# ─── Обробники помилок ────────────────────────────────────────────────────────

@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404


@app.errorhandler(403)
def forbidden(e):
    return render_template('403.html'), 403


@app.errorhandler(429)
def too_many_requests(e):
    return render_template('429.html'), 429


@app.errorhandler(500)
def server_error(e):
    security_logger.error(f'500 error: {e} | path={request.path} | ip={request.remote_addr}')
    return render_template('500.html'), 500


# ─── Маршрути — загальні ──────────────────────────────────────────────────────

@app.route('/')
def index():
    popular_dishes = Dish.query.filter_by(is_active=True).limit(6).all()
    return render_template('index.html', popular_dishes=popular_dishes)


@app.route('/about')
def about():
    return render_template('about.html')


@app.route('/register', methods=['GET', 'POST'])
@limiter.limit("10 per hour")
def register():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = RegistrationForm()
    if form.validate_on_submit():
        user = User(
            username=sanitize_string(form.username.data),
            email=form.email.data.lower().strip()
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        log_security_event('USER_REGISTERED', f'username={user.username}')
        flash('Реєстрація пройшла успішно!', 'success')
        return redirect(url_for('login'))
    return render_template('register.html', form=form)


@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("20 per hour")
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    form = LoginForm()
    ip   = request.remote_addr

    if is_ip_locked(ip):
        mins = get_remaining_lockout(ip)
        flash(f'Забагато невдалих спроб. Спробуйте через {mins} хв.', 'danger')
        return render_template('login.html', form=form, locked=True)

    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data.lower().strip()).first()
        if user and user.check_password(form.password.data):
            if user.is_blocked:
                log_security_event('BLOCKED_LOGIN_ATTEMPT', f'email={user.email}')
                flash('Ваш акаунт заблоковано. Зверніться до адміністратора.', 'danger')
                return render_template('login.html', form=form)
            reset_login_attempts(ip)
            login_user(user)
            log_security_event('LOGIN_SUCCESS', f'email={user.email}')
            next_page = request.args.get('next')
            # Захист від відкритого редіректу
            if next_page and not next_page.startswith('/'):
                next_page = None
            return redirect(next_page or url_for('menu'))
        else:
            record_failed_login(ip)
            remaining = MAX_ATTEMPTS - login_attempts.get(ip, {}).get('count', 0)
            log_security_event('LOGIN_FAILED', f'email={form.email.data}')
            if remaining > 0:
                flash(f'Невірний email або пароль. Залишилось спроб: {remaining}', 'danger')
            else:
                flash(f'Акаунт заблоковано на {LOCKOUT_MIN} хвилин через підозрілу активність.', 'danger')
    return render_template('login.html', form=form)


@app.route('/logout')
@login_required
def logout():
    log_security_event('LOGOUT')
    logout_user()
    session.clear()
    return redirect(url_for('index'))


# ─── Профіль ──────────────────────────────────────────────────────────────────

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    form    = ProfileForm(obj=current_user)
    pw_form = ChangePasswordForm()
    if form.validate_on_submit():
        new_username = sanitize_string(form.username.data)
        new_email    = form.email.data.lower().strip()
        if new_username != current_user.username:
            if User.query.filter_by(username=new_username).first():
                flash("Це ім'я вже зайняте", 'danger')
                return render_template('profile.html', form=form, pw_form=pw_form)
        if new_email != current_user.email:
            if User.query.filter_by(email=new_email).first():
                flash('Цей email вже використовується', 'danger')
                return render_template('profile.html', form=form, pw_form=pw_form)
        current_user.username = new_username
        current_user.email    = new_email
        db.session.commit()
        log_security_event('PROFILE_UPDATED')
        flash('Профіль оновлено!', 'success')
        return redirect(url_for('profile'))
    return render_template('profile.html', form=form, pw_form=pw_form)


@app.route('/change_password', methods=['POST'])
@login_required
@limiter.limit("5 per hour")
def change_password():
    pw_form = ChangePasswordForm()
    if pw_form.validate_on_submit():
        if not current_user.check_password(pw_form.current_password.data):
            log_security_event('WRONG_PASSWORD_CHANGE_ATTEMPT')
            flash('Невірний поточний пароль', 'danger')
        else:
            current_user.set_password(pw_form.new_password.data)
            db.session.commit()
            log_security_event('PASSWORD_CHANGED')
            flash('Пароль успішно змінено!', 'success')
    else:
        flash('Помилка валідації форми', 'danger')
    return redirect(url_for('profile'))


# ─── Меню ─────────────────────────────────────────────────────────────────────

@app.route('/menu')
def menu():
    category = request.args.get('category', 'all')
    allowed  = ('all', 'appetizers', 'main', 'desserts', 'drinks')
    if category not in allowed:
        category = 'all'
    if category == 'all':
        dishes = Dish.query.filter_by(is_active=True).all()
    else:
        dishes = Dish.query.filter_by(category=category, is_active=True).all()
    return render_template('menu.html', dishes=dishes, category=category)


@app.route('/dish/<int:dish_id>', methods=['GET', 'POST'])
def dish_detail(dish_id):
    dish        = Dish.query.get_or_404(dish_id)
    review_form = ReviewForm()
    user_review = None
    if current_user.is_authenticated:
        user_review = Review.query.filter_by(user_id=current_user.id, dish_id=dish_id).first()

    if review_form.validate_on_submit() and current_user.is_authenticated:
        rating  = int(review_form.rating.data)
        comment = sanitize_string(review_form.comment.data or '', 500)
        if rating not in range(1, 6):
            abort(400)
        if user_review:
            user_review.rating  = rating
            user_review.comment = comment
            flash('Відгук оновлено!', 'success')
        else:
            review = Review(user_id=current_user.id, dish_id=dish_id, rating=rating, comment=comment)
            db.session.add(review)
            flash('Дякуємо за відгук!', 'success')
        db.session.commit()
        return redirect(url_for('dish_detail', dish_id=dish_id))

    if user_review and request.method == 'GET':
        review_form.rating.data  = str(user_review.rating)
        review_form.comment.data = user_review.comment

    return render_template('dish_detail.html', dish=dish, review_form=review_form, user_review=user_review)


@app.route('/delete_review/<int:dish_id>', methods=['POST'])
@login_required
def delete_review(dish_id):
    review = Review.query.filter_by(user_id=current_user.id, dish_id=dish_id).first_or_404()
    db.session.delete(review)
    db.session.commit()
    flash('Відгук видалено', 'success')
    return redirect(url_for('dish_detail', dish_id=dish_id))


# ─── QR-код меню ──────────────────────────────────────────────────────────────

@app.route('/qr_menu')
def qr_menu():
    menu_url = request.host_url.rstrip('/') + url_for('menu')
    qr = qrcode.QRCode(version=1, box_size=8, border=4)
    qr.add_data(menu_url)
    qr.make(fit=True)
    img    = qr.make_image(fill_color='#d4af37', back_color='#0a0e27')
    buf    = io.BytesIO()
    img.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()
    return render_template('qr_menu.html', qr_b64=qr_b64, menu_url=menu_url)


# ─── Кошик ────────────────────────────────────────────────────────────────────

@app.route('/cart')
def cart():
    cart_items     = session.get('cart', {})
    total, items   = 0, []
    for dish_id, qty in cart_items.items():
        dish = Dish.query.get(int(dish_id))
        if dish and dish.is_active:
            item_total = dish.price * qty
            total     += item_total
            items.append({'dish': dish, 'quantity': qty, 'subtotal': item_total})
    promo_discount = session.get('promo_discount', 0)
    promo_code     = session.get('promo_code', '')
    total_after    = round(total * (1 - promo_discount / 100), 2) if promo_discount else total
    return render_template('cart.html', items=items, total=total,
                           promo_discount=promo_discount, promo_code=promo_code, total_after=total_after)


@app.route('/add_to_cart/<int:dish_id>', methods=['POST'])
@login_required
def add_to_cart(dish_id):
    dish = Dish.query.get_or_404(dish_id)
    if not dish.is_active:
        flash('Ця страва недоступна', 'danger')
        return redirect(url_for('menu'))
    quantity = min(max(request.form.get('quantity', 1, type=int), 1), 20)
    if 'cart' not in session:
        session['cart'] = {}
    key = str(dish_id)
    session['cart'][key] = session['cart'].get(key, 0) + quantity
    session.modified = True
    flash('Страву додано до кошика!', 'success')
    return redirect(url_for('menu'))


@app.route('/remove_from_cart/<int:dish_id>')
def remove_from_cart(dish_id):
    if 'cart' in session:
        session['cart'].pop(str(dish_id), None)
        session.modified = True
    return redirect(url_for('cart'))


@app.route('/apply_promo', methods=['POST'])
@login_required
@limiter.limit("10 per hour")
def apply_promo():
    code  = sanitize_string(request.form.get('promo_code', '')).upper()[:50]
    promo = PromoCode.query.filter_by(code=code, is_active=True).first()
    if not promo:
        flash('Промокод не знайдено або він недійсний', 'danger')
    elif promo.uses_left is not None and promo.uses_left <= 0:
        flash('Промокод вже вичерпано', 'danger')
    else:
        session['promo_code']     = promo.code
        session['promo_discount'] = promo.discount
        session.modified = True
        flash(f'Промокод застосовано! Знижка {promo.discount:.0f}%', 'success')
    return redirect(url_for('cart'))


@app.route('/remove_promo')
def remove_promo():
    session.pop('promo_code', None)
    session.pop('promo_discount', None)
    session.modified = True
    return redirect(url_for('cart'))


# ─── Оформлення замовлення ────────────────────────────────────────────────────

@app.route('/checkout', methods=['GET', 'POST'])
@login_required
def checkout():
    if request.method == 'POST':
        cart_items = session.get('cart', {})
        if not cart_items:
            flash('Кошик порожній', 'danger')
            return redirect(url_for('cart'))

        total_price, order_items = 0, []
        for dish_id, qty in cart_items.items():
            dish = Dish.query.get(int(dish_id))
            if dish and dish.is_active:
                qty = min(max(int(qty), 1), 20)
                total_price += dish.price * qty
                order_items.append({'dish_id': int(dish_id), 'quantity': qty, 'price': dish.price})

        promo_discount = session.get('promo_discount', 0)
        promo_code     = session.get('promo_code', '')
        if promo_discount:
            total_price = round(total_price * (1 - promo_discount / 100), 2)
            promo = PromoCode.query.filter_by(code=promo_code).first()
            if promo and promo.uses_left is not None:
                promo.uses_left = max(0, promo.uses_left - 1)
                if promo.uses_left == 0:
                    promo.is_active = False

        order = Order(
            user_id=current_user.id, total_price=total_price,
            status='pending', promo_code=promo_code or None, discount=promo_discount
        )
        db.session.add(order)
        db.session.flush()
        for item in order_items:
            db.session.add(OrderItem(order_id=order.id, dish_id=item['dish_id'],
                                     quantity=item['quantity'], price=item['price']))
        db.session.commit()
        session['cart'] = {}
        session.pop('promo_code', None)
        session.pop('promo_discount', None)
        session.modified = True
        return redirect(url_for('payment_select', order_id=order.id))

    cart_items = session.get('cart', {})
    total, items = 0, []
    for dish_id, qty in cart_items.items():
        dish = Dish.query.get(int(dish_id))
        if dish and dish.is_active:
            item_total = dish.price * qty
            total += item_total
            items.append({'dish': dish, 'quantity': qty, 'subtotal': item_total})
    promo_discount = session.get('promo_discount', 0)
    total_after    = round(total * (1 - promo_discount / 100), 2) if promo_discount else total
    return render_template('checkout.html', items=items, total=total,
                           promo_discount=promo_discount, total_after=total_after)

    cart_items = session.get('cart', {})
    total, items = 0, []
    for dish_id, qty in cart_items.items():
        dish = Dish.query.get(int(dish_id))
        if dish and dish.is_active:
            item_total = dish.price * qty
            total += item_total
            items.append({'dish': dish, 'quantity': qty, 'subtotal': item_total})
    promo_discount = session.get('promo_discount', 0)
    total_after    = round(total * (1 - promo_discount / 100), 2) if promo_discount else total
    return render_template('checkout.html', items=items, total=total,
                           promo_discount=promo_discount, total_after=total_after)


# ─── Замовлення ───────────────────────────────────────────────────────────────

@app.route('/orders')
@login_required
def orders():
    user_orders = Order.query.filter_by(user_id=current_user.id).order_by(Order.created_at.desc()).all()
    return render_template('orders.html', orders=user_orders)


@app.route('/order/<int:order_id>')
@login_required
def order_detail(order_id):
    order = Order.query.get_or_404(order_id)
    if order.user_id != current_user.id and not current_user.is_admin:
        log_security_event('UNAUTHORIZED_ORDER_ACCESS', f'order_id={order_id}')
        abort(403)
    return render_template('order_detail.html', order=order)


@app.route('/cancel_order/<int:order_id>', methods=['POST'])
@login_required
def cancel_order(order_id):
    order = Order.query.get_or_404(order_id)
    if order.user_id != current_user.id:
        log_security_event('UNAUTHORIZED_CANCEL_ATTEMPT', f'order_id={order_id}')
        abort(403)
    if order.status == 'confirmed':
        order.status = 'cancelled'
        db.session.commit()
        flash('Замовлення скасовано', 'success')
    return redirect(url_for('orders'))


# ─── Бронювання ───────────────────────────────────────────────────────────────

@app.route('/booking', methods=['GET', 'POST'])
@login_required
@limiter.limit("10 per hour")
def booking():
    form = BookingForm()
    if form.validate_on_submit():
        booking_date = form.booking_date.data
        if not isinstance(booking_date, datetime):
            try:
                booking_date = datetime.fromisoformat(str(booking_date))
            except Exception:
                flash('Неправильна дата бронювання', 'danger')
                return render_template('booking.html', form=form)
        # Забороняємо бронювання в минулому
        if booking_date < datetime.utcnow():
            flash('Не можна бронювати на минулу дату', 'danger')
            return render_template('booking.html', form=form)
        new_booking = Booking(
            user_id=current_user.id,
            booking_date=booking_date,
            guests=form.guests.data,
            notes=sanitize_string(form.notes.data or '', 500),
            status='pending'
        )
        db.session.add(new_booking)
        db.session.commit()
        send_booking_confirmation(current_user, new_booking)
        flash('Запит на бронювання відправлено!', 'success')
        return redirect(url_for('bookings'))
    return render_template('booking.html', form=form)


@app.route('/bookings')
@login_required
def bookings():
    user_bookings = Booking.query.filter_by(user_id=current_user.id).order_by(Booking.booking_date.desc()).all()
    return render_template('bookings.html', bookings=user_bookings)


# ─── Адмін ────────────────────────────────────────────────────────────────────

@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    total_orders   = Order.query.count()
    total_users    = User.query.count()
    total_bookings = Booking.query.count()
    total_revenue  = db.session.query(db.func.sum(Order.total_price)).filter(
                         Order.status != 'cancelled').scalar() or 0

    # ── Продажі за останні 7 днів ──
    today     = datetime.utcnow().date()
    week_days = [(today - timedelta(days=i)) for i in range(6, -1, -1)]
    week_labels, week_revenue, week_orders = [], [], []
    for day in week_days:
        day_start = datetime(day.year, day.month, day.day)
        day_end   = day_start + timedelta(days=1)
        rev = db.session.query(db.func.sum(Order.total_price)).filter(
            Order.created_at >= day_start,
            Order.created_at < day_end,
            Order.status != 'cancelled'
        ).scalar() or 0
        cnt = Order.query.filter(
            Order.created_at >= day_start,
            Order.created_at < day_end
        ).count()
        week_labels.append(day.strftime('%d.%m'))
        week_revenue.append(round(rev, 2))
        week_orders.append(cnt)

    # ── Продажі за останні 12 місяців ──
    month_labels, month_revenue = [], []
    for i in range(11, -1, -1):
        d = today.replace(day=1) - timedelta(days=i * 30)
        m_start = datetime(d.year, d.month, 1)
        if d.month == 12:
            m_end = datetime(d.year + 1, 1, 1)
        else:
            m_end = datetime(d.year, d.month + 1, 1)
        rev = db.session.query(db.func.sum(Order.total_price)).filter(
            Order.created_at >= m_start,
            Order.created_at < m_end,
            Order.status != 'cancelled'
        ).scalar() or 0
        month_labels.append(d.strftime('%b %Y'))
        month_revenue.append(round(rev, 2))

    # ── Топ-5 страв ──
    top_dishes_raw = db.session.query(
        Dish.name,
        db.func.sum(OrderItem.quantity).label('total_qty'),
        db.func.sum(OrderItem.quantity * OrderItem.price).label('total_rev')
    ).join(OrderItem, Dish.id == OrderItem.dish_id)\
     .group_by(Dish.id)\
     .order_by(db.func.sum(OrderItem.quantity).desc())\
     .limit(5).all()
    top_dishes = [{'name': r[0], 'qty': int(r[1]), 'rev': round(r[2], 2)} for r in top_dishes_raw]

    # ── Статус замовлень ──
    order_statuses = {}
    status_labels_ua = {
        'pending': 'В очікуванні', 'confirmed': 'Підтверджено',
        'preparing': 'Готується', 'ready': 'Готово',
        'delivered': 'Доставлено', 'cancelled': 'Скасовано'
    }
    for status, label in status_labels_ua.items():
        cnt = Order.query.filter_by(status=status).count()
        if cnt:
            order_statuses[label] = cnt

    # ── Нові користувачі за 7 днів ──
    new_users_week = User.query.filter(
        User.created_at >= datetime.utcnow() - timedelta(days=7)
    ).count()

    # ── Середній чек ──
    avg_order = db.session.query(db.func.avg(Order.total_price)).filter(
        Order.status != 'cancelled').scalar() or 0

    return render_template('admin_dashboard.html',
        total_orders=total_orders, total_users=total_users,
        total_bookings=total_bookings, total_revenue=round(total_revenue, 2),
        week_labels=json.dumps(week_labels),
        week_revenue=json.dumps(week_revenue),
        week_orders=json.dumps(week_orders),
        month_labels=json.dumps(month_labels),
        month_revenue=json.dumps(month_revenue),
        top_dishes=top_dishes,
        order_statuses=json.dumps(order_statuses),
        new_users_week=new_users_week,
        avg_order=round(avg_order, 2)
    )


@app.route('/admin/dishes', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_dishes():
    form = DishForm()
    if form.validate_on_submit():
        dish = Dish(
            name=sanitize_string(form.name.data),
            description=sanitize_string(form.description.data or '', 1000),
            price=form.price.data,
            category=form.category.data,
            image_url=sanitize_string(form.image_url.data or '', 255)
        )
        db.session.add(dish)
        db.session.commit()
        flash('Страву додано!', 'success')
        return redirect(url_for('admin_dishes'))
    dishes = Dish.query.all()
    return render_template('admin_dishes.html', form=form, dishes=dishes)


@app.route('/admin/orders')
@login_required
@admin_required
def admin_orders():
    all_orders = Order.query.order_by(Order.created_at.desc()).all()
    return render_template('admin_orders.html', orders=all_orders)


@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    all_users = User.query.all()
    return render_template('admin_users.html', users=all_users)


@app.route('/admin/block_user/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def block_user(user_id):
    user = User.query.get_or_404(user_id)
    if user.is_admin:
        return jsonify({'error': 'Не можна заблокувати адміна'}), 403
    user.is_blocked = not user.is_blocked
    db.session.commit()
    log_security_event('USER_BLOCK_TOGGLE', f'user_id={user_id} blocked={user.is_blocked}')
    return jsonify({'success': True, 'is_blocked': user.is_blocked})


@app.route('/admin/bookings')
@login_required
@admin_required
def admin_bookings():
    all_bookings = Booking.query.order_by(Booking.booking_date.desc()).all()
    return render_template('admin_bookings.html', bookings=all_bookings)


@app.route('/admin/promos', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_promos():
    form = PromoForm()
    if form.validate_on_submit():
        code = sanitize_string(form.code.data).upper()[:50]
        if PromoCode.query.filter_by(code=code).first():
            flash('Такий промокод вже існує', 'danger')
        else:
            uses  = form.uses_left.data if form.uses_left.data and form.uses_left.data > 0 else None
            promo = PromoCode(code=code, discount=form.discount.data, uses_left=uses)
            db.session.add(promo)
            db.session.commit()
            flash(f'Промокод {code} створено!', 'success')
        return redirect(url_for('admin_promos'))
    promos = PromoCode.query.order_by(PromoCode.created_at.desc()).all()
    return render_template('admin_promos.html', form=form, promos=promos)


@app.route('/admin/toggle_promo/<int:promo_id>', methods=['POST'])
@login_required
@admin_required
def toggle_promo(promo_id):
    promo = PromoCode.query.get_or_404(promo_id)
    promo.is_active = not promo.is_active
    db.session.commit()
    return jsonify({'success': True, 'is_active': promo.is_active})


@app.route('/admin/delete_promo/<int:promo_id>', methods=['POST'])
@login_required
@admin_required
def delete_promo(promo_id):
    promo = PromoCode.query.get_or_404(promo_id)
    db.session.delete(promo)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/admin/update_order/<int:order_id>/<status>', methods=['POST'])
@login_required
@admin_required
def update_order_status(order_id, status):
    valid = ('pending', 'confirmed', 'preparing', 'ready', 'delivered', 'cancelled')
    if status not in valid:
        return jsonify({'error': 'Невірний статус'}), 400
    order = Order.query.get_or_404(order_id)
    order.status = status
    db.session.commit()
    if status in ('confirmed', 'preparing', 'ready', 'delivered', 'cancelled'):
        send_order_status_update(order.customer, order)
    return jsonify({'success': True})


@app.route('/admin/update_booking/<int:booking_id>/<status>', methods=['POST'])
@login_required
@admin_required
def update_booking_status(booking_id, status):
    valid = ('pending', 'confirmed', 'cancelled')
    if status not in valid:
        return jsonify({'error': 'Невірний статус'}), 400
    booking = Booking.query.get_or_404(booking_id)
    booking.status = status
    db.session.commit()
    if status in ('confirmed', 'cancelled'):
        send_booking_status_update(booking.customer, booking)
    return jsonify({'success': True})


@app.route('/admin/delete_dish/<int:dish_id>', methods=['POST'])
@login_required
@admin_required
def delete_dish(dish_id):
    dish = Dish.query.get_or_404(dish_id)
    db.session.delete(dish)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/admin/delete_review/<int:review_id>', methods=['POST'])
@login_required
@admin_required
def admin_delete_review(review_id):
    review = Review.query.get_or_404(review_id)
    db.session.delete(review)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/admin/email')
@login_required
@admin_required
def admin_email():
    return render_template('admin_email.html',
                           mail_enabled=MAIL_ENABLED,
                           mail_sender=app.config.get('MAIL_DEFAULT_SENDER', ''))


@app.route('/admin/send_test_email', methods=['POST'])
@login_required
@admin_required
def admin_send_test_email():
    if not MAIL_ENABLED:
        flash('Email не налаштовано', 'danger')
        return redirect(url_for('admin_email'))
    html = f"""
    <div style="font-family:sans-serif;max-width:500px;background:#0a0e27;color:#e0e0e0;border-radius:12px;padding:24px;">
      <h2 style="color:#d4af37;">✦ Bella Cucina — Тестовий лист</h2>
      <p>Вітаємо, <strong>{current_user.username}</strong>!</p>
      <p>Email сповіщення налаштовано успішно. ✅</p>
    </div>"""
    send_email('Тестовий лист — Bella Cucina', [current_user.email], html)
    flash(f'Тестовий лист надіслано на {current_user.email}', 'success')
    return redirect(url_for('admin_email'))


# ─── Контекст для AI чату (меню доступне у всіх шаблонах) ───────────────────

@app.context_processor
def inject_menu_context():
    try:
        dishes = Dish.query.filter_by(is_active=True).all()
        menu_context = [
            {'name': d.name, 'description': d.description,
             'price': d.price, 'category': d.category}
            for d in dishes
        ]
    except Exception:
        menu_context = []
    return {'menu_context': menu_context}


# ─── AI Чат-асистент (локальний бот, без API) ────────────────────────────────

def local_bot_reply(message: str, dishes: list) -> str:
    msg = message.lower().strip()

    by_cat = {}
    for d in dishes:
        by_cat.setdefault(d.category, []).append(d)

    appetizers = by_cat.get('appetizers', [])
    mains      = by_cat.get('main', [])
    desserts   = by_cat.get('desserts', [])
    drinks     = by_cat.get('drinks', [])
    all_dishes = dishes

    # ── Привітання ──
    if any(w in msg for w in ['привіт', 'hello', 'hi', 'добрий', 'вітаю', 'салют', 'hey']):
        return 'Вітаю! 👋 Радий бачити вас у Bella Cucina! Я шеф-асистент і знаю все про наше меню.\nЩо вас цікавить — закуски, основні страви, десерти чи напої? Або просто скажіть що любите — і я підберу ідеально! 🍽️'

    # ── Рекомендація загальна ──
    if any(w in msg for w in ['рекоменд', 'порад', 'що обрати', 'що замовити', 'що взяти', 'найкращ', 'топ', 'хіт']):
        top_main = mains[0] if mains else None
        top_des  = desserts[0] if desserts else None
        reply = '🌟 Мої особисті рекомендації:\n'
        if top_main: reply += f'• **{top_main.name}** (${top_main.price:.2f}) — {top_main.description[:60]}...\n'
        if top_des:  reply += f'• **{top_des.name}** (${top_des.price:.2f}) — {top_des.description[:60]}...\n'
        reply += '\nОбидві страви — абсолютні фаворити наших гостей! 😋'
        return reply

    # ── Популярне ──
    if any(w in msg for w in ['популярн', 'часто замовл', 'улюблен', 'фаворит']):
        picks = (mains[:1] + desserts[:1] + drinks[:1])
        lines = '\n'.join(f'• {d.name} (${d.price:.2f})' for d in picks)
        return f'Найпопулярніші страви Bella Cucina:\n{lines}\n\nСаме їх замовляють найчастіше! 🏆'

    # ── Закуски ──
    if any(w in msg for w in ['закуск', 'аперитив', 'почат', 'стартер', 'перед основн']):
        lines = '\n'.join(f'• {d.name} — ${d.price:.2f}: {d.description[:55]}...' for d in appetizers)
        return f'Наші закуски:\n{lines}\n\n🥗 Ідеально щоб розпочати вечерю! Брускета — легка та свіжа, Кальмарі — для любителів морепродуктів.'

    # ── Основні страви ──
    if any(w in msg for w in ['основн', 'головн', 'паста', 'карбонара', 'лазань', 'різото', 'оссо', 'м\'яс', 'друге']):
        lines = '\n'.join(f'• {d.name} — ${d.price:.2f}: {d.description[:55]}...' for d in mains)
        return f'Основні страви:\n{lines}\n\n🍝 Всі готуються щодня зі свіжих інгредієнтів. Карбонара — класика Риму, Різото — для справжніх гурманів!'

    # ── Десерти ──
    if any(w in msg for w in ['десерт', 'солодк', 'тірамісу', 'панна', 'торт', 'морозив', 'солодощ', 'після їжі']):
        lines = '\n'.join(f'• {d.name} — ${d.price:.2f}: {d.description[:55]}...' for d in desserts)
        return f'Десерти:\n{lines}\n\n🍮 Тірамісу — наша гордість за класичним рецептом. Панна Котта тане в роті! Обидва варті уваги 😍'

    # ── Напої ──
    if any(w in msg for w in ['напій', 'напо', 'кава', 'еспресо', 'вино', 'просекко', 'пити', 'чай', 'випити']):
        lines = '\n'.join(f'• {d.name} — ${d.price:.2f}: {d.description[:55]}...' for d in drinks)
        return f'Напої:\n{lines}\n\n☕🍷 Еспресо — справжній італійський, міцний та ароматний. Просекко — ідеально для святкування!'

    # ── Вегетаріанське ──
    if any(w in msg for w in ['вегетар', 'без м\'яс', 'овочев', 'веган', 'пост']):
        veg = [d for d in all_dishes if any(w in d.description.lower() for w in ['гриб', 'овоч', 'помідор', 'базилік', 'ванільн', 'вершков', 'лісов'])]
        if veg:
            lines = '\n'.join(f'• {d.name} (${d.price:.2f})' for d in veg[:4])
            return f'Страви без м\'яса:\n{lines}\n\n🌿 Різото з грибами — особливо рекомендую! Ситно, смачно і повністю вегетаріанське.'
        return 'У нас є чудові варіанти без м\'яса — Брускета, Різото з грибами та всі десерти. Смачно і легко! 🌿'

    # ── Дешево ──
    if any(w in msg for w in ['дешев', 'бюджет', 'недорог', 'економ', 'мало грош']):
        cheap = sorted(all_dishes, key=lambda d: d.price)[:4]
        lines = '\n'.join(f'• {d.name} — ${d.price:.2f}' for d in cheap)
        return f'Найдоступніші страви:\n{lines}\n\n💰 Відмінна якість за розумною ціною! Смак Італії для кожного.'

    # ── Дорого / преміум ──
    if any(w in msg for w in ['дорог', 'преміум', 'розкіш', 'особлив', 'святкув', 'дата']):
        prem = sorted(all_dishes, key=lambda d: d.price, reverse=True)[:3]
        lines = '\n'.join(f'• {d.name} — ${d.price:.2f}' for d in prem)
        return f'Преміум страви для особливих моментів:\n{lines}\n\n✨ Оссо Буко — справжній кулінарний шедевр по-міланськи!'

    # ── Романтика / побачення ──
    if any(w in msg for w in ['романтик', 'побачен', 'дівчин', 'хлопц', 'закоха', 'валентин', 'річниц']):
        return '💕 Для романтичного вечора рекомендую:\n• Оссо Буко або Різото з грибами на основне\n• Панна Котта на десерт\n• Просекко до страв\n\nЗабронюйте столик заздалегідь — підберемо найкраще місце! 🕯️'

    # ── Компанія / група ──
    if any(w in msg for w in ['компані', 'груп', 'друз', 'колег', 'корпорат', 'велик']):
        return '👥 Для великої компанії рекомендуємо:\n• Різні закуски на стіл — Брускета та Кальмарі\n• Асорті основних страв\n• Просекко для настрою\n\nЗабронюйте столик заздалегідь — ми підготуємось! 📅'

    # ── Діти ──
    if any(w in msg for w in ['дит', 'дитин', 'малюк', 'дітям']):
        return '👶 Для дітей чудово підійдуть:\n• Спагеті Карбонара (без алкоголю)\n• Тірамісу або Панна Котта на десерт\n• Еспресо замінимо на сік\n\nМи завжди раді маленьким гостям! 🍝'

    # ── Швидко / поспішаю ──
    if any(w in msg for w in ['швидко', 'поспішаю', 'мало час', 'терміново', 'швидк']):
        fast = [d for d in all_dishes if d.price < 15][:2]
        lines = '\n'.join(f'• {d.name} (${d.price:.2f})' for d in fast)
        return f'Якщо поспішаєте — рекомендую:\n{lines}\n\n⚡ Готується швидко і дуже смачно!'

    # ── Бронювання ──
    if any(w in msg for w in ['брон', 'столик', 'місц', 'резерв', 'забронюва', 'сісти']):
        return '📅 Забронювати столик легко!\n\n1. Натисніть **Бронювання** в меню\n2. Оберіть дату і час\n3. Вкажіть кількість гостей\n4. Підтвердження прийде на email\n\nЗазвичай підтверджуємо протягом 10 хвилин! ✅'

    # ── Адреса / час роботи ──
    if any(w in msg for w in ['адрес', 'де знаход', 'як доїхат', 'де ви', 'місцезнаход']):
        return '📍 Ми знаходимось:\n**вул. Хрещатик, 1, Київ**\n\nЗупинка метро: Хрещатик (синя лінія)\nПаркінг: є поряд з рестораном 🚗'

    if any(w in msg for w in ['години', 'час роботи', 'коли відчин', 'коли закрив', 'графік', 'розклад']):
        return '🕐 Режим роботи:\n**Щодня: 10:00 – 23:00**\n\nОстаннє замовлення приймаємо о 22:30. Чекаємо на вас! 😊'

    # ── Ціна / меню ──
    if any(w in msg for w in ['скільки', 'ціна', 'вартість', 'прайс', 'price']):
        cheap = min(all_dishes, key=lambda d: d.price) if all_dishes else None
        pricey = max(all_dishes, key=lambda d: d.price) if all_dishes else None
        if cheap and pricey:
            return f'💰 Ціни у нас від **${cheap.price:.2f}** ({cheap.name}) до **${pricey.price:.2f}** ({pricey.name}).\n\nСередній чек на одну особу: ~$25–35. Перегляньте повне меню для деталей!'

    # ── Меню ──
    if any(w in msg for w in ['меню', 'страви', 'список', 'що є', 'що маєте', 'асортимент']):
        return f'🍽️ Наше меню:\n• Закуски: {len(appetizers)} страви\n• Основні: {len(mains)} страви\n• Десерти: {len(desserts)} страви\n• Напої: {len(drinks)} позиції\n\nВсього {len(all_dishes)} позицій! Запитайте про будь-яку категорію або перейдіть до розділу Меню.'

    # ── Дякую ──
    if any(w in msg for w in ['дякую', 'дякуємо', 'спасибі', 'спасибо', 'thanks', 'дяки']):
        return 'Будь ласка! 😊 Завжди радий допомогти.\nПриємного апетиту та чудового вечора у Bella Cucina! 🍷✨'

    # ── Добре / окей ──
    if any(w in msg for w in ['добре', 'окей', 'ок', 'зрозуміло', 'ясно', 'чудово']):
        return 'Чудово! 😊 Якщо виникнуть ще питання — я тут. Можете також переглянути повне меню на сайті! 🍽️'

    # ── Дефолт ──
    import random
    tips = [
        'Що порекомендуєте на вечерю?',
        'Є вегетаріанські страви?',
        'Які десерти маєте?',
        'Скільки коштує?',
        'Як забронювати столик?',
    ]
    tip = random.choice(tips)
    return f'Гарне питання! 🤔 Я спеціалізуюся на стравах та інформації про ресторан.\nСпробуйте запитати: **"{tip}"** або оберіть тему нижче!'


@app.route('/ai_chat', methods=['POST'])
@limiter.limit("30 per minute")
def ai_chat():
    import time
    data = request.get_json(silent=True) or {}
    user_message = sanitize_string(data.get('message', ''), 500)

    if not user_message:
        return jsonify({'reply': 'Напишіть своє питання!'})

    try:
        dishes = Dish.query.filter_by(is_active=True).all()
    except Exception:
        dishes = []

    reply = local_bot_reply(user_message, dishes)

    # Штучна затримка — імітація "друкування" (0.6–1.2 сек залежно від довжини)
    delay = min(0.4 + len(reply) * 0.003, 1.4)
    time.sleep(delay)

    return jsonify({'reply': reply})

@app.route('/admin/export_orders')
@login_required
@admin_required
def export_orders():
    from flask import make_response
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Замовлення'

    # Стилі
    header_font    = Font(bold=True, color='1A1A1A', size=11)
    header_fill    = PatternFill('solid', fgColor='D4AF37')
    center_align   = Alignment(horizontal='center', vertical='center')
    wrap_align     = Alignment(wrap_text=True, vertical='center')
    alt_fill       = PatternFill('solid', fgColor='1A1F3A')

    headers = ['№', 'Клієнт', 'Email', 'Страви', 'Сума ($)', 'Знижка (%)', 'Промокод', 'Статус', 'Дата']
    widths  = [6, 20, 28, 40, 12, 12, 14, 16, 18]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    status_ua = {
        'pending': 'В очікуванні', 'confirmed': 'Підтверджено',
        'preparing': 'Готується',  'ready': 'Готово',
        'delivered': 'Доставлено', 'cancelled': 'Скасовано'
    }

    orders = Order.query.order_by(Order.created_at.desc()).all()
    for i, order in enumerate(orders, 2):
        dishes_text = '; '.join(
            f'{item.dish.name} ×{item.quantity}' for item in order.items
        )
        row_data = [
            order.id,
            order.customer.username,
            order.customer.email,
            dishes_text,
            round(order.total_price, 2),
            int(order.discount) if order.discount else 0,
            order.promo_code or '—',
            status_ua.get(order.status, order.status),
            order.created_at.strftime('%d.%m.%Y %H:%M')
        ]
        for col, value in enumerate(row_data, 1):
            cell           = ws.cell(row=i, column=col, value=value)
            cell.alignment = wrap_align if col == 4 else center_align
            if i % 2 == 0:
                cell.fill = alt_fill

        ws.row_dimensions[i].height = 18

    # Підсумковий рядок
    last = len(orders) + 2
    ws.cell(row=last, column=4, value='РАЗОМ:').font = Font(bold=True, color='D4AF37')
    total = sum(o.total_price for o in orders if o.status != 'cancelled')
    ws.cell(row=last, column=5, value=round(total, 2)).font = Font(bold=True, color='D4AF37')

    # Лист «Топ страв»
    ws2 = wb.create_sheet('Топ страв')
    ws2.append(['Страва', 'Замовлень (шт)', 'Дохід ($)'])
    for col_idx, w in enumerate([30, 18, 16], 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = w
    for cell in ws2[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align

    top = db.session.query(
        Dish.name,
        db.func.sum(OrderItem.quantity).label('qty'),
        db.func.sum(OrderItem.quantity * OrderItem.price).label('rev')
    ).join(OrderItem).group_by(Dish.id).order_by(db.func.sum(OrderItem.quantity).desc()).all()

    for i, (name, qty, rev) in enumerate(top, 2):
        ws2.cell(row=i, column=1, value=name).alignment = wrap_align
        ws2.cell(row=i, column=2, value=int(qty)).alignment = center_align
        ws2.cell(row=i, column=3, value=round(rev, 2)).alignment = center_align
        if i % 2 == 0:
            for c in range(1, 4):
                ws2.cell(row=i, column=c).fill = alt_fill

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'bella_cucina_orders_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.xlsx'
    response = make_response(buf.read())
    response.headers['Content-Type']        = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'
    return response


# ─── Оплата — вибір методу ────────────────────────────────────────────────────

@app.route('/payment/<int:order_id>')
@login_required
def payment_select(order_id):
    order = Order.query.get_or_404(order_id)
    if order.user_id != current_user.id:
        abort(403)
    if order.status not in ('pending',):
        flash('Це замовлення вже оброблено.', 'info')
        return redirect(url_for('order_detail', order_id=order.id))
    return render_template('payment_select.html',
                           order=order,
                           stripe_enabled=STRIPE_ENABLED,
                           liqpay_enabled=LIQPAY_ENABLED,
                           stripe_public_key=STRIPE_PUBLIC_KEY)


# ─── Готівка ──────────────────────────────────────────────────────────────────

@app.route('/payment/<int:order_id>/cash', methods=['POST'])
@login_required
def payment_cash(order_id):
    order = Order.query.get_or_404(order_id)
    if order.user_id != current_user.id:
        abort(403)
    payment = Payment(
        order_id=order.id, user_id=current_user.id,
        amount=order.total_price, currency='UAH',
        provider='cash', status='pending'
    )
    db.session.add(payment)
    order.status = 'confirmed'
    db.session.commit()
    send_order_confirmation(current_user, order)
    flash('Замовлення оформлено! Оплата готівкою при отриманні.', 'success')
    return redirect(url_for('order_detail', order_id=order.id))


# ─── Stripe ───────────────────────────────────────────────────────────────────

@app.route('/payment/<int:order_id>/stripe/create', methods=['POST'])
@login_required
def stripe_create_intent(order_id):
    order = Order.query.get_or_404(order_id)
    if order.user_id != current_user.id:
        abort(403)
    stripe = get_stripe()
    if not stripe:
        return jsonify({'error': 'Stripe не налаштовано'}), 503
    try:
        amount_cents = int(round(order.total_price * 100))
        intent = stripe.PaymentIntent.create(
            amount=amount_cents,
            currency='usd',
            metadata={'order_id': order.id, 'user_id': current_user.id},
            description=f'Bella Cucina — Замовлення #{order.id}'
        )
        # Зберігаємо або оновлюємо Payment
        payment = order.payment
        if not payment:
            payment = Payment(order_id=order.id, user_id=current_user.id,
                              amount=order.total_price, currency='USD', provider='stripe')
            db.session.add(payment)
        payment.external_id = intent.id
        payment.status = 'pending'
        db.session.commit()
        return jsonify({'client_secret': intent.client_secret})
    except Exception as e:
        security_logger.error(f'[STRIPE_ERROR] {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/payment/<int:order_id>/stripe/success', methods=['POST'])
@login_required
def stripe_payment_success(order_id):
    order = Order.query.get_or_404(order_id)
    if order.user_id != current_user.id:
        abort(403)
    payment = order.payment
    if payment:
        payment.status = 'paid'
    order.status = 'confirmed'
    db.session.commit()
    send_order_confirmation(current_user, order)
    flash('Оплата пройшла успішно! Дякуємо!', 'success')
    return jsonify({'success': True, 'redirect': url_for('order_detail', order_id=order.id)})


@app.route('/webhook/stripe', methods=['POST'])
@csrf.exempt
def stripe_webhook():
    stripe = get_stripe()
    if not stripe or not STRIPE_WEBHOOK_SECRET:
        abort(503)
    payload   = request.get_data()
    sig       = request.headers.get('Stripe-Signature', '')
    try:
        event = stripe.Webhook.construct_event(payload, sig, STRIPE_WEBHOOK_SECRET)
    except Exception as e:
        security_logger.error(f'[STRIPE_WEBHOOK_ERROR] {e}')
        return jsonify({'error': str(e)}), 400

    if event['type'] == 'payment_intent.succeeded':
        intent   = event['data']['object']
        order_id = intent.get('metadata', {}).get('order_id')
        if order_id:
            order = Order.query.get(int(order_id))
            if order:
                if order.payment:
                    order.payment.status = 'paid'
                order.status = 'confirmed'
                db.session.commit()
    return jsonify({'received': True})


# ─── LiqPay ───────────────────────────────────────────────────────────────────

@app.route('/payment/<int:order_id>/liqpay')
@login_required
def liqpay_pay(order_id):
    order = Order.query.get_or_404(order_id)
    if order.user_id != current_user.id:
        abort(403)
    if not LIQPAY_ENABLED:
        flash('LiqPay не налаштовано', 'danger')
        return redirect(url_for('payment_select', order_id=order_id))

    liqpay_order_id = f'bella_{order.id}_{int(datetime.utcnow().timestamp())}'
    data = {
        'version':     '3',
        'public_key':  LIQPAY_PUBLIC_KEY,
        'action':      'pay',
        'amount':      str(round(order.total_price, 2)),
        'currency':    'UAH',
        'description': f'Bella Cucina — Замовлення #{order.id}',
        'order_id':    liqpay_order_id,
        'result_url':  request.host_url + f'payment/{order.id}/liqpay/result',
        'server_url':  request.host_url + 'webhook/liqpay',
    }
    data_b64, signature = liqpay_encode(data)

    # Зберігаємо Payment
    payment = order.payment
    if not payment:
        payment = Payment(order_id=order.id, user_id=current_user.id,
                          amount=order.total_price, currency='UAH', provider='liqpay')
        db.session.add(payment)
    payment.external_id = liqpay_order_id
    payment.status = 'pending'
    db.session.commit()

    return render_template('liqpay_pay.html', order=order,
                           data_b64=data_b64, signature=signature)


@app.route('/payment/<int:order_id>/liqpay/result')
@login_required
def liqpay_result(order_id):
    order = Order.query.get_or_404(order_id)
    return redirect(url_for('order_detail', order_id=order.id))


@app.route('/webhook/liqpay', methods=['POST'])
@csrf.exempt
def liqpay_webhook():
    if not LIQPAY_ENABLED:
        abort(503)
    data_b64  = request.form.get('data', '')
    signature = request.form.get('signature', '')
    decoded   = liqpay_decode(data_b64, signature)
    if not decoded:
        security_logger.error('[LIQPAY_WEBHOOK] Невірний підпис')
        abort(400)

    liqpay_order_id = decoded.get('order_id', '')
    status          = decoded.get('status', '')
    # Знаходимо Payment за external_id
    payment = Payment.query.filter_by(external_id=liqpay_order_id).first()
    if payment:
        if status in ('success', 'sandbox'):
            payment.status        = 'paid'
            payment.order.status  = 'confirmed'
            db.session.commit()
            with app.app_context():
                send_order_confirmation(payment.payer, payment.order)
        elif status in ('failure', 'error'):
            payment.status = 'failed'
            db.session.commit()
    return 'OK'


# ─── Повернення коштів ────────────────────────────────────────────────────────

@app.route('/admin/refund/<int:payment_id>', methods=['POST'])
@login_required
@admin_required
def admin_refund(payment_id):
    payment = Payment.query.get_or_404(payment_id)
    reason  = sanitize_string(request.form.get('reason', 'Повернення коштів'), 500)

    if payment.status != 'paid':
        return jsonify({'error': 'Можна повернути лише оплачені платежі'}), 400

    if payment.provider == 'stripe':
        stripe = get_stripe()
        if stripe and payment.external_id:
            try:
                stripe.Refund.create(payment_intent=payment.external_id)
            except Exception as e:
                return jsonify({'error': str(e)}), 500

    elif payment.provider == 'liqpay' and LIQPAY_ENABLED:
        data = {
            'version':    '3',
            'public_key': LIQPAY_PUBLIC_KEY,
            'action':     'refund',
            'order_id':   payment.external_id,
            'amount':     str(round(payment.amount, 2)),
            'currency':   payment.currency,
        }
        data_b64, sig = liqpay_encode(data)
        import urllib.request, urllib.parse
        params = urllib.parse.urlencode({'data': data_b64, 'signature': sig}).encode()
        urllib.request.urlopen('https://www.liqpay.ua/api/request', params)

    payment.status        = 'refunded'
    payment.refund_reason = reason
    payment.order.status  = 'cancelled'
    db.session.commit()
    log_security_event('REFUND', f'payment_id={payment_id} order_id={payment.order_id}')
    return jsonify({'success': True})


# ─── Історія платежів (адмін) ─────────────────────────────────────────────────

@app.route('/admin/payments')
@login_required
@admin_required
def admin_payments():
    payments = Payment.query.order_by(Payment.created_at.desc()).all()
    return render_template('admin_payments.html', payments=payments)


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=False, host='0.0.0.0', port=5000)